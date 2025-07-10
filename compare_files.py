import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import yagmail
import tempfile

# Inject custom background color that actually works
st.markdown("""
    <style>
    html, body, [data-testid="stAppViewContainer"] {
        background-color: #f0f8ff;
    }
    </style>
""", unsafe_allow_html=True)



st.set_page_config(page_title="File Comparison Tool", layout="centered")
st.title("🔍 Compare Two Files (CSV or Excel)")
st.write("Upload two files (CSV or Excel). The app will highlight numeric values that are not present in the other file.")

# File uploaders
data1 = st.file_uploader("Upload first file", type=["csv", "xlsx", "xls"], key="file1")
data2 = st.file_uploader("Upload second file", type=["csv", "xlsx", "xls"], key="file2")

# Helper to read file
def read_data(uploaded_file):
    if uploaded_file.name.lower().endswith('.csv'):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)

# Highlighting logic
def highlight_diffs_in_files(df1, df2):
    red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    def extract_numeric_values(df):
        values = set()
        for col in df.columns:
            for val in df[col]:
                try:
                    num = float(val)
                    values.add(round(num, 2))
                except:
                    continue
        return values

    values1 = extract_numeric_values(df1)
    values2 = extract_numeric_values(df2)

    def save_with_highlight(df, other_values):
        wb = Workbook()
        ws = wb.active
        ws.append(df.columns.tolist())
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    continue
                try:
                    num = float(val)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if round(num, 2) not in other_values:
                        cell.fill = red
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=val)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    out1 = save_with_highlight(df1, values2)
    out2 = save_with_highlight(df2, values1)
    return out1, out2

# Use session state to persist results until new files are uploaded
if 'outputs' not in st.session_state:
    st.session_state['outputs'] = None
    st.session_state['last_files'] = (None, None)

# Detect if new files are uploaded and clear outputs if so
current_files = (data1.name if data1 else None, data2.name if data2 else None)
if current_files != st.session_state['last_files']:
    st.session_state['outputs'] = None
    st.session_state['last_files'] = current_files

if data1 and data2:
    try:
        df1 = read_data(data1)
        df2 = read_data(data2)
        st.success("Files uploaded! Click below to compare.")
        if st.button("Compare Files"):
            out1, out2 = highlight_diffs_in_files(df1, df2)
            st.session_state['outputs'] = (out1, out2)
        if st.session_state['outputs']:
            out1, out2 = st.session_state['outputs']
            st.write("### Download Results:")
            st.download_button(
                label="Download Highlighted File 1",
                data=out1,
                file_name="file1_highlighted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.download_button(
                label="Download Highlighted File 2",
                data=out2,
                file_name="file2_highlighted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            # Email input and send button appear only after download links
            email = st.text_input("Enter your email to receive the results (optional):")
            if st.button("Send Results"):
                if email:
                    try:
                        sender_email = "kfirslon@gmail.com"
                        app_password = "hwqcvquvxhjcldat"  # 16 chars, no spaces
                        yag = yagmail.SMTP(sender_email, app_password)
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp1, \
                             tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp2:
                            tmp1.write(out1.getvalue())
                            tmp1.flush()
                            tmp2.write(out2.getvalue())
                            tmp2.flush()
                            yag.send(
                                to=email,
                                subject="Your File Comparison Results",
                                contents="Attached are the highlighted comparison results for your files.",
                                attachments=[tmp1.name, tmp2.name]
                            )
                        st.success(f"Results sent to {email}!")
                    except Exception as e:
                        st.error(f"Failed to send email: {e}")
                else:
                    st.warning("Please enter an email address to send the results.")
    except Exception as e:
        st.error(f"Error processing files: {e}")
